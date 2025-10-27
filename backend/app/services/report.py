from typing import Dict, Any, Optional
from datetime import datetime
from io import BytesIO
import pandas as pd

from app.database.raw_query_executor import RawQueryExecutor
from app.core.logging import get_logger

logger = get_logger(__name__)


class ReportService:
    """Service for generating Excel reports from database data."""

    def __init__(self):
        """Initialize the report service with a RawQueryExecutor."""
        self.executor = RawQueryExecutor()

    def generate_comprehensive_report(self) -> BytesIO:
        """
        Generate a comprehensive Excel report with multiple sheets.
        """
        try:
            logger.info("Starting comprehensive report generation")

            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1
                df_calculations = self._get_calculation_history_df()
                df_calculations.to_excel(
                    writer,
                    sheet_name='Calculation History',
                    index=False
                )
                logger.info(f"Added Calculation History sheet with {len(df_calculations)} rows")

                # Sheet 2
                df_results = self._get_detailed_results_df()
                df_results.to_excel(
                    writer,
                    sheet_name='Detailed Results',
                    index=False
                )
                logger.info(f"Added Detailed Results sheet with {len(df_results)} rows")

                # Sheet 3
                df_party_summary = self._get_party_summary_df()
                df_party_summary.to_excel(
                    writer,
                    sheet_name='Party Performance',
                    index=False
                )
                logger.info(f"Added Party Performance sheet with {len(df_party_summary)} rows")

                # Sheet 4
                df_submissions = self._get_voting_submissions_df()
                df_submissions.to_excel(
                    writer,
                    sheet_name='Voting Submissions',
                    index=False
                )
                logger.info(f"Added Voting Submissions sheet with {len(df_submissions)} rows")

                # Sheet 5
                df_stats = self._get_statistics_df()
                df_stats.to_excel(
                    writer,
                    sheet_name='Statistics',
                    index=False
                )
                logger.info("Added Statistics sheet")

                self._apply_formatting(writer)

            output.seek(0)

            logger.info("Comprehensive report generated successfully")
            return output

        except Exception as e:
            logger.error(f"Failed to generate report: {e}", exc_info=True)
            raise RuntimeError(f"Report generation failed: {e}")

    def _get_calculation_history_df(self) -> pd.DataFrame:
        """
        Get calculation history with summary information.
        """
        sql = """
            SELECT
                c.id as calculation_id,
                c.timestamp,
                c.total_seats,
                c.total_votes,
                COUNT(cr.id) as parties_count,
                SUM(CASE WHEN cr.seats > 0 THEN 1 ELSE 0 END) as parties_with_seats,
                MAX(cr.seats) as max_seats_won,
                ROUND(AVG(cr.seats)::numeric, 2) as avg_seats_per_party
            FROM calculations c
            LEFT JOIN calculation_results cr ON c.id = cr.calculation_id
            GROUP BY c.id, c.timestamp, c.total_seats, c.total_votes
            ORDER BY c.timestamp DESC
        """

        results = self.executor.execute_query(sql, return_format="dict")

        if not results:
            return pd.DataFrame(columns=[
                'Calculation ID', 'Date', 'Time', 'Total Seats',
                'Total Votes', 'Parties Count', 'Parties with Seats',
                'Max Seats Won', 'Avg Seats per Party'
            ])

        df = pd.DataFrame(results)

        df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize(None)
        df['date'] = df['timestamp'].dt.date
        df['time'] = df['timestamp'].dt.time

        df = df[[
            'calculation_id', 'date', 'time', 'total_seats', 'total_votes',
            'parties_count', 'parties_with_seats', 'max_seats_won',
            'avg_seats_per_party'
        ]]

        df.columns = [
            'Calculation ID', 'Date', 'Time', 'Total Seats', 'Total Votes',
            'Parties Count', 'Parties with Seats', 'Max Seats Won',
            'Avg Seats per Party'
        ]

        return df

    def _get_detailed_results_df(self) -> pd.DataFrame:
        """
        Get detailed results for all calculations with party information.
        """
        sql = """
            SELECT
                c.id as calculation_id,
                c.timestamp,
                p.name as party_name,
                cr.votes,
                cr.seats,
                c.total_votes,
                c.total_seats,
                ROUND((cr.votes::numeric / NULLIF(c.total_votes, 0) * 100), 2) as vote_percentage,
                ROUND((cr.seats::numeric / NULLIF(c.total_seats, 0) * 100), 2) as seat_percentage
            FROM calculations c
            INNER JOIN calculation_results cr ON c.id = cr.calculation_id
            INNER JOIN parties p ON cr.party_id = p.id
            ORDER BY c.timestamp DESC, cr.seats DESC, cr.votes DESC
        """

        results = self.executor.execute_query(sql, return_format="dict")

        if not results:
            return pd.DataFrame(columns=[
                'Calculation ID', 'Timestamp', 'Party', 'Votes', 'Seats',
                'Total Votes', 'Total Seats', 'Vote %', 'Seat %'
            ])

        df = pd.DataFrame(results)
        # Remove timezone info for Excel compatibility
        df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize(None)

        df.columns = [
            'Calculation ID', 'Timestamp', 'Party', 'Votes', 'Seats',
            'Total Votes', 'Total Seats', 'Vote %', 'Seat %'
        ]

        return df

    def _get_party_summary_df(self) -> pd.DataFrame:
        """
        Get aggregated performance summary for each party across all calculations.
        """
        sql = """
            SELECT
                p.name as party_name,
                COUNT(DISTINCT c.id) as total_calculations,
                SUM(cr.votes) as total_votes,
                SUM(cr.seats) as total_seats,
                ROUND(AVG(cr.votes)::numeric, 2) as avg_votes_per_calc,
                ROUND(AVG(cr.seats)::numeric, 2) as avg_seats_per_calc,
                MAX(cr.seats) as max_seats_in_single_calc,
                MIN(cr.seats) as min_seats_in_single_calc,
                SUM(CASE WHEN cr.seats > 0 THEN 1 ELSE 0 END) as times_won_seats,
                ROUND((SUM(CASE WHEN cr.seats > 0 THEN 1 ELSE 0 END)::numeric /
                      NULLIF(COUNT(DISTINCT c.id), 0) * 100), 2) as win_rate_percentage
            FROM parties p
            LEFT JOIN calculation_results cr ON p.id = cr.party_id
            LEFT JOIN calculations c ON cr.calculation_id = c.id
            GROUP BY p.id, p.name
            ORDER BY total_seats DESC NULLS LAST, total_votes DESC NULLS LAST
        """

        results = self.executor.execute_query(sql, return_format="dict")

        if not results:
            return pd.DataFrame(columns=[
                'Party', 'Total Calculations', 'Total Votes', 'Total Seats',
                'Avg Votes per Calc', 'Avg Seats per Calc', 'Max Seats',
                'Min Seats', 'Times Won Seats', 'Win Rate %'
            ])

        df = pd.DataFrame(results)

        df.columns = [
            'Party', 'Total Calculations', 'Total Votes', 'Total Seats',
            'Avg Votes per Calc', 'Avg Seats per Calc', 'Max Seats',
            'Min Seats', 'Times Won Seats', 'Win Rate %'
        ]

        # Fill NaN values with 0 for parties with no calculations
        df = df.fillna(0)

        return df

    def _get_voting_submissions_df(self) -> pd.DataFrame:
        """
        Get all voting submissions with party information.
        """
        sql = """
            SELECT
                vs.id as submission_id,
                vs.submitted_at,
                p.name as party_name,
                vs.votes,
                DATE(vs.submitted_at) as submission_date,
                EXTRACT(HOUR FROM vs.submitted_at) as submission_hour
            FROM voting_submissions vs
            INNER JOIN parties p ON vs.party_id = p.id
            ORDER BY vs.submitted_at DESC
        """

        results = self.executor.execute_query(sql, return_format="dict")

        if not results:
            return pd.DataFrame(columns=[
                'Submission ID', 'Submitted At', 'Party', 'Votes',
                'Submission Date', 'Hour of Day'
            ])

        df = pd.DataFrame(results)
        # Remove timezone info for Excel compatibility
        df['submitted_at'] = pd.to_datetime(df['submitted_at']).dt.tz_localize(None)

        df.columns = [
            'Submission ID', 'Submitted At', 'Party', 'Votes',
            'Submission Date', 'Hour of Day'
        ]

        return df

    def _get_statistics_df(self) -> pd.DataFrame:
        """
        Get overall statistics summary.
        """
        sql = """
            SELECT
                'Total Calculations' as metric,
                COUNT(*)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'Total Parties' as metric,
                COUNT(*)::text as value
            FROM parties

            UNION ALL

            SELECT
                'Total Voting Submissions' as metric,
                COUNT(*)::text as value
            FROM voting_submissions

            UNION ALL

            SELECT
                'Total Votes Across All Calculations' as metric,
                SUM(total_votes)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'Total Seats Allocated' as metric,
                SUM(total_seats)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'Average Seats per Calculation' as metric,
                ROUND(AVG(total_seats)::numeric, 2)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'First Calculation Date' as metric,
                MIN(timestamp)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'Last Calculation Date' as metric,
                MAX(timestamp)::text as value
            FROM calculations

            UNION ALL

            SELECT
                'Most Active Party (by votes)' as metric,
                COALESCE(
                    (SELECT p.name
                     FROM parties p
                     LEFT JOIN calculation_results cr ON p.id = cr.party_id
                     GROUP BY p.id, p.name
                     ORDER BY SUM(cr.votes) DESC NULLS LAST
                     LIMIT 1),
                    'N/A'
                ) as value

            UNION ALL

            SELECT
                'Most Successful Party (by seats)' as metric,
                COALESCE(
                    (SELECT p.name
                     FROM parties p
                     LEFT JOIN calculation_results cr ON p.id = cr.party_id
                     GROUP BY p.id, p.name
                     ORDER BY SUM(cr.seats) DESC NULLS LAST
                     LIMIT 1),
                    'N/A'
                ) as value
        """

        results = self.executor.execute_query(sql, return_format="dict")

        if not results:
            return pd.DataFrame(columns=['Metric', 'Value'])

        df = pd.DataFrame(results)
        df.columns = ['Metric', 'Value']

        timestamp_row = pd.DataFrame({
            'Metric': ['Report Generated At'],
            'Value': [datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')]
        })
        df = pd.concat([timestamp_row, df], ignore_index=True)

        return df

    def _apply_formatting(self, writer: pd.ExcelWriter) -> None:
        """
        Apply formatting to Excel sheets.
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                header_fill = PatternFill(start_color="366092",
                                         end_color="366092",
                                         fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center",
                                              vertical="center")

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                worksheet.freeze_panes = 'A2'

            logger.info("Applied formatting to all sheets")

        except Exception as e:
            logger.warning(f"Could not apply full formatting: {e}")

    def generate_custom_report(
        self,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        party_filter: Optional[str] = None
    ) -> BytesIO:
        """
        Generate a custom filtered report.
        """
        try:
            logger.info(f"Generating custom report with filters: "
                       f"start={start_date}, end={end_date}, party={party_filter}")

            where_clauses = []
            params = {}

            if start_date:
                where_clauses.append("c.timestamp >= :start_date")
                params['start_date'] = start_date

            if end_date:
                where_clauses.append("c.timestamp <= :end_date")
                params['end_date'] = end_date

            if party_filter:
                where_clauses.append("p.name = :party_name")
                params['party_name'] = party_filter

            where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"

            sql = f"""
                SELECT
                    c.id as calculation_id,
                    c.timestamp,
                    p.name as party_name,
                    cr.votes,
                    cr.seats,
                    c.total_votes,
                    c.total_seats
                FROM calculations c
                INNER JOIN calculation_results cr ON c.id = cr.calculation_id
                INNER JOIN parties p ON cr.party_id = p.id
                WHERE {where_clause}
                ORDER BY c.timestamp DESC, cr.seats DESC
            """

            results = self.executor.execute_query(sql, params=params, return_format="dict")

            if not results:
                logger.warning("No data found for custom report filters")
                results = []

            df = pd.DataFrame(results)

            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Custom Report', index=False)
                self._apply_formatting(writer)

            output.seek(0)
            logger.info(f"Custom report generated with {len(df)} rows")
            return output

        except Exception as e:
            logger.error(f"Failed to generate custom report: {e}", exc_info=True)
            raise RuntimeError(f"Custom report generation failed: {e}")


# Singleton instance
_report_service = None


def get_report_service() -> ReportService:
    """
    Get or create the singleton ReportService instance.
    """
    global _report_service
    if _report_service is None:
        _report_service = ReportService()
    return _report_service
